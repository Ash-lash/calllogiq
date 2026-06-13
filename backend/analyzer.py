import os
import sys
import json
import argparse
from datetime import datetime, timedelta
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define duration splits in order
DURATION_RANGES = [
    "0 sec (Not Connected)",
    "1 - 10 secs",
    "11 - 20 secs",
    "21 - 30 secs",
    "31 - 45 secs",
    "46 - 60 secs",
    "1 min 1 sec - 1 min 30 secs",
    "1 min 31 secs - 2 mins",
    "2 mins 1 sec - 3 mins",
    "3 mins 1 sec - 5 mins",
    "Above 5 mins"
]

def parse_duration(dur_str):
    if not dur_str:
        return 0
    parts = dur_str.strip().split(':')
    try:
        if len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
        elif len(parts) == 2:
            return int(parts[0]) * 60 + int(parts[1])
        elif len(parts) == 1:
            return int(parts[0])
    except ValueError:
        return 0
    return 0

def get_duration_range(seconds):
    if seconds == 0:
        return "0 sec (Not Connected)"
    elif 1 <= seconds <= 10:
        return "1 - 10 secs"
    elif 11 <= seconds <= 20:
        return "11 - 20 secs"
    elif 21 <= seconds <= 30:
        return "21 - 30 secs"
    elif 31 <= seconds <= 45:
        return "31 - 45 secs"
    elif 46 <= seconds <= 60:
        return "46 - 60 secs"
    elif 61 <= seconds <= 90:
        return "1 min 1 sec - 1 min 30 secs"
    elif 91 <= seconds <= 120:
        return "1 min 31 secs - 2 mins"
    elif 121 <= seconds <= 180:
        return "2 mins 1 sec - 3 mins"
    elif 181 <= seconds <= 300:
        return "3 mins 1 sec - 5 mins"
    else:
        return "Above 5 mins"

def format_seconds(seconds):
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    return f"{h:02d}:{m:02d}:{s:02d}"

def analyze_pdf(pdf_path, username):
    all_calls = []
    
    # Extract data from PDF
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    # Validate row shape and content
                    if len(row) < 5:
                        continue
                    name, phone, time_str, duration_str, call_type = row[0], row[1], row[2], row[3], row[4]
                    
                    # Skip header row or empty entries
                    if name == 'Name' and phone == 'Phone Number':
                        continue
                    if not time_str or not call_type:
                        continue
                        
                    # Standardize names and numbers
                    name = (name or '').strip().replace('\n', ' ')
                    phone = (phone or '').strip()
                    time_str = (time_str or '').strip()
                    duration_str = (duration_str or '').strip()
                    call_type = (call_type or '').strip()
                    
                    if not call_type in ['Dialed', 'Received', 'Missed']:
                        continue
                        
                    duration_secs = parse_duration(duration_str)
                    
                    # Parse time
                    try:
                        dt = datetime.strptime(time_str, "%H:%M %d-%m-%Y")
                    except ValueError:
                        try:
                            # Try parsing fallback
                            dt = datetime.strptime(time_str, "%H:%M %d-%m")
                            # Set current year or placeholder
                            dt = dt.replace(year=2026)
                        except ValueError:
                            # Skip if time format is unparseable
                            continue
                            
                    all_calls.append({
                        'name': name or phone,
                        'phone': phone,
                        'time': dt,
                        'time_str': time_str,
                        'duration_str': duration_str,
                        'duration_secs': duration_secs,
                        'type': call_type
                    })
                    
    if not all_calls:
        raise ValueError("No call log records found in PDF")
        
    # Sort calls chronologically
    all_calls.sort(key=lambda x: x['time'])
    
    # Find most common call date
    dates = [c['time'].strftime("%d-%m-%Y") for c in all_calls]
    call_date_str = max(set(dates), key=dates.count)
    call_date = datetime.strptime(call_date_str, "%d-%m-%Y")
    
    # Calculate counts
    dialed_calls = [c for c in all_calls if c['type'] == 'Dialed']
    incoming_calls = [c for c in all_calls if c['type'] == 'Received']
    missed_calls = [c for c in all_calls if c['type'] == 'Missed']
    
    total_dialed = len(dialed_calls)
    total_incoming = len(incoming_calls)
    total_missed = len(missed_calls)
    grand_total = total_dialed + total_incoming + total_missed
    
    # Active Workday & Idle Hours Analysis
    workday_start = all_calls[0]['time']
    workday_end = all_calls[-1]['time'] + timedelta(seconds=all_calls[-1]['duration_secs'])
    total_workday_secs = int((workday_end - workday_start).total_seconds())
    
    # Calculate talk time
    total_talk_secs = sum(c['duration_secs'] for c in all_calls)
    
    # Calculate Idle Gaps (> 15 minutes)
    idle_gaps = []
    total_idle_secs = 0
    for i in range(len(all_calls) - 1):
        call_end_time = all_calls[i]['time'] + timedelta(seconds=all_calls[i]['duration_secs'])
        next_call_start_time = all_calls[i+1]['time']
        
        gap_secs = int((next_call_start_time - call_end_time).total_seconds())
        if gap_secs > 900:  # 15 minutes
            idle_gaps.append({
                'start': call_end_time.strftime("%H:%M"),
                'end': next_call_start_time.strftime("%H:%M"),
                'duration_secs': gap_secs,
                'duration_str': format_seconds(gap_secs)
            })
            total_idle_secs += gap_secs
            
    # Duration Splitups
    dialed_splits = {r: 0 for r in DURATION_RANGES}
    incoming_splits = {r: 0 for r in DURATION_RANGES}
    
    for c in dialed_calls:
        r = get_duration_range(c['duration_secs'])
        dialed_splits[r] += 1
        
    for c in incoming_calls:
        r = get_duration_range(c['duration_secs'])
        incoming_splits[r] += 1
        
    # Hourly distribution
    hourly_distribution = {}
    for h in range(24):
        hourly_distribution[f"{h:02d}:00"] = 0
    for c in all_calls:
        hour_str = c['time'].strftime("%H:00")
        hourly_distribution[hour_str] += 1
    # Filter out hours with 0 calls to keep JSON clean, but keep a full view
    hourly_distribution = {k: v for k, v in hourly_distribution.items() if v > 0}
    
    # Connected Calls
    connected_dialed = sum(1 for c in dialed_calls if c['duration_secs'] > 0)
    connected_incoming = sum(1 for c in incoming_calls if c['duration_secs'] > 0)
    
    analysis = {
        'username': username,
        'call_date': call_date.strftime("%d %b %Y"),
        'call_date_filename': call_date.strftime("%d%b%Y"),
        'summary': {
            'total_dialed': total_dialed,
            'total_incoming': total_incoming,
            'total_missed': total_missed,
            'grand_total': grand_total,
            'connected_dialed': connected_dialed,
            'connected_incoming': connected_incoming,
            'talk_time_secs': total_talk_secs,
            'talk_time_str': format_seconds(total_talk_secs),
            'avg_duration_secs': int(total_talk_secs / (connected_dialed + connected_incoming)) if (connected_dialed + connected_incoming) > 0 else 0,
            'workday_start': workday_start.strftime("%H:%M"),
            'workday_end': workday_end.strftime("%H:%M"),
            'workday_span_secs': total_workday_secs,
            'workday_span_str': format_seconds(total_workday_secs),
            'total_idle_secs': total_idle_secs,
            'total_idle_str': format_seconds(total_idle_secs),
            'idle_gaps_count': len(idle_gaps),
            'idle_gaps': idle_gaps,
            'hourly_distribution': hourly_distribution,
            'dialed_splits': dialed_splits,
            'incoming_splits': incoming_splits
        },
        'calls': [
            {
                'name': c['name'],
                'phone': c['phone'],
                'time': c['time'].strftime("%Y-%m-%d %H:%M:%S"),
                'duration': c['duration_str'],
                'duration_secs': c['duration_secs'],
                'type': c['type']
            } for c in all_calls
        ]
    }
    
    return analysis

def generate_excel(analysis, output_path):
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    ws = wb.create_sheet(title='Summary')
    
    # Grid lines visible
    ws.views.sheetView[0].showGridLines = True
    
    # Set column widths
    ws.column_dimensions['A'].width = 32.0
    ws.column_dimensions['B'].width = 14.0
    ws.column_dimensions['C'].width = 16.0
    ws.column_dimensions['D'].width = 28.0
    
    # Styles
    font_family = "Arial"
    
    # Fonts
    font_title = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    font_section = Font(name=font_family, size=12, bold=True, color="FFFFFF")
    font_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    font_normal = Font(name=font_family, size=11, bold=False, color="000000")
    font_bold = Font(name=font_family, size=11, bold=True, color="000000")
    
    # Fills
    fill_title = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_section = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    fill_header = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    fill_summary_row = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft green
    fill_grand_total = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft yellow
    fill_zebra_light = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid") # Soft grey-blue
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Borders
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    thick_bottom_side = Side(border_style="medium", color="000000")
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_total = Border(top=Side(border_style="thin", color="000000"), bottom=Side(border_style="double", color="000000"))
    
    # Helpers to apply styles to merged ranges
    def style_merged_range(ws, cell_range, font, fill, alignment):
        first_cell = ws[cell_range.split(':')[0]]
        first_cell.font = font
        first_cell.fill = fill
        first_cell.alignment = alignment
        
        # Apply border/fill to all cells in range
        rows = ws[cell_range]
        for r in rows:
            for cell in r:
                cell.fill = fill
                cell.border = border_cell
                
    # --- ROW 1: Title ---
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 35
    title_text = f"TELECALLER CALL LOG ANALYSIS - {analysis['call_date']}"
    ws['A1'] = title_text
    style_merged_range(ws, 'A1:D1', font_title, fill_title, align_center)
    
    # --- ROW 3: Overall Call Summary ---
    ws.merge_cells('A3:D3')
    ws.row_dimensions[3].height = 24
    ws['A3'] = "OVERALL CALL SUMMARY"
    style_merged_range(ws, 'A3:D3', font_section, fill_section, align_center)
    
    # --- ROW 4: Table Headers ---
    ws.row_dimensions[4].height = 20
    headers_summary = ["Category", "Count", "% of Total", "Remarks"]
    for idx, h in enumerate(headers_summary, 1):
        cell = ws.cell(row=4, column=idx)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if idx in [2, 3] else align_left
        cell.border = border_cell
        
    # --- ROWS 5-7: Summary Data ---
    sum_data = [
        ("Total Dialled Calls", analysis['summary']['total_dialed'], "Calls made by telecaller"),
        ("Total Incoming Calls", analysis['summary']['total_incoming'], "Calls received"),
        ("Total Missed Calls", analysis['summary']['total_missed'], "Missed / unanswered")
    ]
    
    for offset, (cat, count, remark) in enumerate(sum_data, 5):
        ws.row_dimensions[offset].height = 20
        # Category
        cell_cat = ws.cell(row=offset, column=1, value=cat)
        cell_cat.alignment = align_left
        # Count
        cell_cnt = ws.cell(row=offset, column=2, value=count)
        cell_cnt.alignment = align_right
        cell_cnt.number_format = '#,##0'
        # % of Total (Formula)
        cell_pct = ws.cell(row=offset, column=3, value=f"=B{offset}/$B$8")
        cell_pct.alignment = align_right
        cell_pct.number_format = '0.00%'
        # Remarks
        cell_rem = ws.cell(row=offset, column=4, value=remark)
        cell_rem.alignment = align_left
        
        for idx in range(1, 5):
            c = ws.cell(row=offset, column=idx)
            c.font = font_normal
            c.fill = fill_summary_row
            c.border = border_cell
            
    # --- ROW 8: Grand Total ---
    ws.row_dimensions[8].height = 20
    ws.cell(row=8, column=1, value="GRAND TOTAL").alignment = align_left
    ws.cell(row=8, column=2, value="=SUM(B5:B7)").alignment = align_right
    ws.cell(row=8, column=2).number_format = '#,##0'
    ws.cell(row=8, column=3, value="=SUM(C5:C7)").alignment = align_right
    ws.cell(row=8, column=3).number_format = '0.00%'
    ws.cell(row=8, column=4, value="All calls combined").alignment = align_left
    
    for idx in range(1, 5):
        c = ws.cell(row=8, column=idx)
        c.font = font_bold
        c.fill = fill_grand_total
        c.border = border_cell
        
    # --- ROW 10: Dialled Split Up Header ---
    ws.merge_cells('A10:D10')
    ws.row_dimensions[10].height = 24
    ws['A10'] = "DIALLED CALLS - DURATION SPLIT UP"
    style_merged_range(ws, 'A10:D10', font_section, fill_section, align_center)
    
    # --- ROW 11: Dialled Headers ---
    ws.row_dimensions[11].height = 20
    headers_dialled = ["Duration Range", "No. of Calls", "% of Dialled", "Cumulative %"]
    for idx, h in enumerate(headers_dialled, 1):
        cell = ws.cell(row=11, column=idx)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if idx in [2, 3, 4] else align_left
        cell.border = border_cell
        
    # --- ROWS 12-22: Dialled split data ---
    for offset, dur_range in enumerate(DURATION_RANGES, 12):
        ws.row_dimensions[offset].height = 20
        # Alternating fills
        row_fill = fill_zebra_light if offset % 2 == 0 else fill_white
        
        ws.cell(row=offset, column=1, value=dur_range).alignment = align_left
        ws.cell(row=offset, column=2, value=analysis['summary']['dialed_splits'].get(dur_range, 0)).alignment = align_right
        ws.cell(row=offset, column=2).number_format = '#,##0'
        
        # % of Dialled
        ws.cell(row=offset, column=3, value=f"=B{offset}/$B$23").alignment = align_right
        ws.cell(row=offset, column=3).number_format = '0.00%'
        
        # Cumulative %
        if offset == 12:
            ws.cell(row=offset, column=4, value="=C12").alignment = align_right
        else:
            ws.cell(row=offset, column=4, value=f"=D{offset-1}+C{offset}").alignment = align_right
        ws.cell(row=offset, column=4).number_format = '0.00%'
        
        for idx in range(1, 5):
            c = ws.cell(row=offset, column=idx)
            c.font = font_normal
            c.fill = row_fill
            c.border = border_cell
            
    # --- ROW 23: Total Dialled ---
    ws.row_dimensions[23].height = 20
    ws.cell(row=23, column=1, value="TOTAL DIALLED").alignment = align_left
    ws.cell(row=23, column=2, value="=SUM(B12:B22)").alignment = align_right
    ws.cell(row=23, column=2).number_format = '#,##0'
    ws.cell(row=23, column=3, value="=SUM(C12:C22)").alignment = align_right
    ws.cell(row=23, column=3).number_format = '0.00%'
    ws.cell(row=23, column=4, value="")
    
    for idx in range(1, 5):
        c = ws.cell(row=23, column=idx)
        c.font = font_bold
        c.fill = fill_grand_total
        c.border = border_cell
        
    # --- ROW 25: Incoming Split Up Header ---
    ws.merge_cells('A25:D25')
    ws.row_dimensions[25].height = 24
    ws['A25'] = "INCOMING CALLS - DURATION SPLIT UP"
    style_merged_range(ws, 'A25:D25', font_section, fill_section, align_center)
    
    # --- ROW 26: Incoming Headers ---
    ws.row_dimensions[26].height = 20
    headers_incoming = ["Duration Range", "No. of Calls", "% of Incoming", ""]
    # Ah! In sample merged ranges, it had MergedCellRange A25:D25, but row 26 was:
    # Row 26: ['Duration Range', 'No. of Calls', '% of Incoming', None] (Not merged!)
    # Let's write them individually:
    
    # First unmerge A26:D26 just in case we merged it:
    # ws.unmerge_cells('A26:D26')
    
    for idx, h in enumerate(headers_incoming, 1):
        cell = ws.cell(row=26, column=idx)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if idx in [2, 3] else align_left
        cell.border = border_cell
        
    # --- ROWS 27-37: Incoming split data ---
    for offset, dur_range in enumerate(DURATION_RANGES, 27):
        ws.row_dimensions[offset].height = 20
        # Alternating fills
        row_fill = fill_zebra_light if offset % 2 == 1 else fill_white # Alternating offset starting at 27
        
        ws.cell(row=offset, column=1, value=dur_range).alignment = align_left
        ws.cell(row=offset, column=2, value=analysis['summary']['incoming_splits'].get(dur_range, 0)).alignment = align_right
        ws.cell(row=offset, column=2).number_format = '#,##0'
        
        # % of Incoming
        ws.cell(row=offset, column=3, value=f"=B{offset}/$B$38").alignment = align_right
        ws.cell(row=offset, column=3).number_format = '0.00%'
        
        ws.cell(row=offset, column=4, value="")
        
        for idx in range(1, 5):
            c = ws.cell(row=offset, column=idx)
            c.font = font_normal
            c.fill = row_fill
            c.border = border_cell
            
    # --- ROW 38: Total Incoming ---
    ws.row_dimensions[38].height = 20
    ws.cell(row=38, column=1, value="TOTAL INCOMING").alignment = align_left
    ws.cell(row=38, column=2, value="=SUM(B27:B37)").alignment = align_right
    ws.cell(row=38, column=2).number_format = '#,##0'
    ws.cell(row=38, column=3, value="=SUM(C27:C37)").alignment = align_right
    ws.cell(row=38, column=3).number_format = '0.00%'
    ws.cell(row=38, column=4, value="")
    
    for idx in range(1, 5):
        c = ws.cell(row=38, column=idx)
        c.font = font_bold
        c.fill = fill_grand_total
        c.border = border_cell

    # Save Excel workbook
    wb.save(output_path)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Analyze daily call log PDF and output statistics/Excel.")
    parser.add_argument('--pdf', required=True, help="Path to input PDF file")
    parser.add_argument('--user', required=True, help="Name of the user/employee")
    parser.add_argument('--out', required=True, help="Path to write Excel sheet")
    
    args = parser.parse_args()
    
    try:
        if not os.path.exists(args.pdf):
            print(json.dumps({'error': f"File not found: {args.pdf}"}))
            sys.exit(1)
            
        result = analyze_pdf(args.pdf, args.user)
        generate_excel(result, args.out)
        
        # Return summary and paths as JSON
        print(json.dumps(result, default=str))
        
    except Exception as e:
        import traceback
        err_msg = str(e)
        # Detailed traceback to stderr
        sys.stderr.write(traceback.format_exc())
        print(json.dumps({'error': err_msg}))
        sys.exit(1)
