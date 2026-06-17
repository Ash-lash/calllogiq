"""
generate_user_report.py
Generates an aggregate Excel report for a single user from JSON data passed via stdin.
Produces:
  - Sheet "Summary"  : one row per date (dialed, incoming, missed, talk time, etc.)
  - Sheet per date   : full call-by-call detail for that day
"""

import sys
import json
import argparse
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found – installing...", file=sys.stderr)
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ── Colour palette ─────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1A1A2E")
ACCENT_FILL   = PatternFill("solid", fgColor="16213E")
ALT_ROW_FILL  = PatternFill("solid", fgColor="F0F4FF")
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
TITLE_FONT    = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
BODY_FONT     = Font(name="Calibri", size=10)
BOLD_FONT     = Font(bold=True, name="Calibri", size=10)
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN_BORDER   = Border(
    left=Side(style="thin",   color="D0D7E3"),
    right=Side(style="thin",  color="D0D7E3"),
    top=Side(style="thin",    color="D0D7E3"),
    bottom=Side(style="thin", color="D0D7E3"),
)

def hdr_row(ws, row, cols, values):
    """Write a styled header row."""
    for c, val in zip(cols, values):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER

def data_row(ws, row, cols, values, alt=False):
    """Write a styled data row."""
    fill = ALT_ROW_FILL if alt else WHITE_FILL
    for c, val in zip(cols, values):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font      = BODY_FONT
        cell.fill      = fill
        cell.border    = THIN_BORDER
        cell.alignment = LEFT

def auto_width(ws):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

def title_block(ws, title, subtitle, merge_cols=10):
    """Insert a two-row title block at the top."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=merge_cols)

    t_cell = ws.cell(row=1, column=1, value=title)
    t_cell.font      = TITLE_FONT
    t_cell.fill      = HEADER_FILL
    t_cell.alignment = CENTER

    s_cell = ws.cell(row=2, column=1, value=subtitle)
    s_cell.font      = Font(bold=False, color="FFFFFF", name="Calibri", size=10)
    s_cell.fill      = ACCENT_FILL
    s_cell.alignment = CENTER

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    return 3   # next writable row


def build_summary_sheet(ws, user_name, logs):
    """Sheet 1 – Summary across all dates."""
    subtitle = f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}  |  Employee: {user_name}"
    next_row = title_block(ws, f"Call Log Summary – {user_name}", subtitle, merge_cols=10)

    cols     = list(range(1, 11))
    hdr_vals = ["Date", "Dialed", "Incoming", "Missed", "Total Calls",
                "Talk Time", "Idle Time", "Clock In", "Clock Out", "Span"]
    hdr_row(ws, next_row, cols, hdr_vals)
    next_row += 1

    # Sort logs by callDate (newest first)
    sorted_logs = sorted(logs, key=lambda l: l.get("createdAt", ""), reverse=True)

    for idx, log in enumerate(sorted_logs):
        s = log.get("summary", {})
        vals = [
            log.get("callDate", ""),
            s.get("total_dialed",   0),
            s.get("total_incoming", 0),
            s.get("total_missed",   0),
            s.get("grand_total",    0),
            s.get("talk_time_str",  ""),
            s.get("total_idle_str", ""),
            s.get("workday_start",  ""),
            s.get("workday_end",    ""),
            s.get("workday_span_str", ""),
        ]
        data_row(ws, next_row, cols, vals, alt=(idx % 2 == 1))
        next_row += 1

    if not sorted_logs:
        ws.cell(row=next_row, column=1, value="No call logs found.").font = BODY_FONT

    auto_width(ws)


def build_date_sheet(ws, log):
    """One sheet per date – full call-by-call detail."""
    call_date = log.get("callDate", "Unknown")
    s         = log.get("summary",  {})
    calls     = log.get("calls",    [])

    subtitle = (f"Date: {call_date}  |  "
                f"Total: {s.get('grand_total', 0)}  |  "
                f"Talk Time: {s.get('talk_time_str', '')}  |  "
                f"Span: {s.get('workday_span_str', '')}")

    merge_cols = 7
    next_row = title_block(ws, f"Call Detail – {call_date}", subtitle, merge_cols=merge_cols)

    cols     = list(range(1, merge_cols + 1))
    hdr_vals = ["#", "Contact / Number", "Type", "Start Time", "Duration", "Status", "Notes"]
    hdr_row(ws, next_row, cols, hdr_vals)
    next_row += 1

    if calls:
        for idx, call in enumerate(calls):
            vals = [
                idx + 1,
                call.get("contact") or call.get("number") or call.get("name", ""),
                call.get("type",     call.get("call_type", "")),
                call.get("time",     call.get("start_time", call.get("timestamp", ""))),
                call.get("duration", call.get("duration_str", "")),
                call.get("status",   ""),
                call.get("notes",    call.get("note", "")),
            ]
            data_row(ws, next_row, cols, vals, alt=(idx % 2 == 1))
            next_row += 1
    else:
        ws.cell(row=next_row, column=1,
                value="No individual call records stored for this date.").font = BODY_FONT

    # Totals row
    next_row += 1
    totals_vals = ["TOTALS", "", "",
                   f"Dialed: {s.get('total_dialed', 0)}",
                   f"Incoming: {s.get('total_incoming', 0)}",
                   f"Missed: {s.get('total_missed', 0)}",
                   f"Total: {s.get('grand_total', 0)}"]
    for c, val in zip(cols, totals_vals):
        cell = ws.cell(row=next_row, column=c, value=val)
        cell.font   = BOLD_FONT
        cell.fill   = ACCENT_FILL
        cell.border = THIN_BORDER
        cell.font   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)

    auto_width(ws)


def sanitize_sheet_name(name):
    """Excel sheet names cannot exceed 31 chars or contain certain chars."""
    bad = ['\\', '/', '*', '[', ']', ':', '?']
    for ch in bad:
        name = name.replace(ch, ' ')
    return name[:31]


def main():
    parser = argparse.ArgumentParser(description="Generate aggregate user Excel report.")
    parser.add_argument("--output", required=True, help="Output .xlsx file path")
    args = parser.parse_args()

    # Read JSON from stdin
    raw = sys.stdin.read().strip()
    if not raw:
        print("No data received on stdin.", file=sys.stderr)
        sys.exit(1)

    data = json.loads(raw)
    user_name = data.get("userName", "Employee")
    logs      = data.get("logs",     [])

    wb = openpyxl.Workbook()

    # Sheet 1 – Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    build_summary_sheet(ws_summary, user_name, logs)

    # One sheet per date
    sorted_logs = sorted(logs, key=lambda l: l.get("createdAt", ""), reverse=True)
    for log in sorted_logs:
        sheet_name = sanitize_sheet_name(log.get("callDate", "Unknown"))
        ws = wb.create_sheet(title=sheet_name)
        build_date_sheet(ws, log)

    wb.save(args.output)
    print(f"Report saved to {args.output}", file=sys.stderr)
    sys.exit(0)


if __name__ == "__main__":
    main()
