"""
generate_attendance_report.py
Generates a styled Excel report for a single user or all employees' attendance history from JSON data passed via stdin.
"""

import sys
import json
import argparse
import re
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found – installing...", file=sys.stderr)
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ── Colour palette ─────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1A1A2E")
ACCENT_FILL   = PatternFill("solid", fgColor="16213E")
ALT_ROW_FILL  = PatternFill("solid", fgColor="F0F4FF")
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")

# Status specific fills
PRESENT_FILL  = PatternFill("solid", fgColor="E2F0D9") # Soft green
ABSENT_FILL   = PatternFill("solid", fgColor="FCE4D6") # Soft red
HOLIDAY_FILL  = PatternFill("solid", fgColor="FFF2CC") # Soft yellow
OVERTIME_FILL = PatternFill("solid", fgColor="DDEBF7") # Soft blue

HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
TITLE_FONT    = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
BODY_FONT     = Font(name="Calibri", size=10)
BOLD_FONT     = Font(bold=True, name="Calibri", size=10)
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)

THIN_BORDER   = Border(
    left=Side(style="thin",   color="D0D7E3"),
    right=Side(style="thin",  color="D0D7E3"),
    top=Side(style="thin",    color="D0D7E3"),
    bottom=Side(style="thin", color="D0D7E3"),
)

def auto_width(ws):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

def build_single_user_sheet(ws, data):
    ws.views.sheetView[0].showGridLines = True
    user_name = data.get("userName", "Employee")
    domain = data.get("domain", "-")
    branch = data.get("branch", "-")
    reg_date = data.get("registrationDate", "-")
    summary = data.get("summary", {})
    history = data.get("history", [])

    # Title Block
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    t_cell = ws.cell(row=1, column=1, value=f"ATTENDANCE REPORT – {user_name.upper()}")
    t_cell.font = TITLE_FONT
    t_cell.fill = HEADER_FILL
    t_cell.alignment = CENTER

    subtitle = f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}  |  Domain: {domain}  |  Branch: {branch}  |  Joined: {reg_date}"
    s_cell = ws.cell(row=2, column=1, value=subtitle)
    s_cell.font = Font(bold=False, color="FFFFFF", name="Calibri", size=10)
    s_cell.fill = ACCENT_FILL
    s_cell.alignment = CENTER

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

    # Summary Table
    ws.cell(row=4, column=1, value="METRIC").font = Font(bold=True, name="Calibri", size=10)
    ws.cell(row=4, column=2, value="VALUE").font = Font(bold=True, name="Calibri", size=10)

    metrics = [
        ("Total Working Days", summary.get("workingDays", 0)),
        ("Present Days", summary.get("presentDays", 0)),
        ("Absent Days", summary.get("absentDays", 0)),
        ("Holidays", summary.get("holidays", 0)),
        ("Overtime Days", summary.get("overtimeDays", 0)),
    ]

    pres = summary.get("presentDays", 0)
    total_work = summary.get("workingDays", 0)
    rate = f"{round((pres / total_work) * 100, 1)}%" if total_work > 0 else "0%"
    metrics.append(("Attendance Rate", rate))

    for idx, (m_label, m_val) in enumerate(metrics):
        r = 5 + idx
        c1 = ws.cell(row=r, column=1, value=m_label)
        c2 = ws.cell(row=r, column=2, value=m_val)
        c1.font = BOLD_FONT if m_label == "Attendance Rate" else BODY_FONT
        c2.font = BOLD_FONT if m_label == "Attendance Rate" else BODY_FONT
        c1.border = THIN_BORDER
        c2.border = THIN_BORDER
        c1.alignment = LEFT
        c2.alignment = CENTER
        if m_label == "Attendance Rate":
            c1.fill = ALT_ROW_FILL
            c2.fill = ALT_ROW_FILL

    # History Table Headers
    start_history_row = 13
    ws.cell(row=start_history_row-1, column=1, value="ATTENDANCE HISTORY").font = Font(bold=True, name="Calibri", size=12)

    headers = ["Date", "Status", "Clock In", "Clock Out", "Duration", "Net Work Hours", "Talk Time", "Total Calls"]
    for c_idx, h in enumerate(headers):
        cell = ws.cell(row=start_history_row, column=c_idx+1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # History Table Data
    for idx, day in enumerate(history):
        r = start_history_row + 1 + idx
        status = day.get("status", "Absent")
        
        status_fill = WHITE_FILL
        if status == "Present": status_fill = PRESENT_FILL
        elif status == "Absent": status_fill = ABSENT_FILL
        elif status == "Holiday": status_fill = HOLIDAY_FILL
        elif status == "Overtime": status_fill = OVERTIME_FILL

        row_fill = ALT_ROW_FILL if idx % 2 == 1 else WHITE_FILL

        vals = [
            day.get("date", ""),
            status,
            day.get("arrival", "-"),
            day.get("departure", "-"),
            day.get("duration", "-"),
            day.get("netWorkHours", "-"),
            day.get("talkTime", "-"),
            day.get("calls", 0)
        ]

        for c_idx, val in enumerate(vals):
            cell = ws.cell(row=r, column=c_idx+1, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER if c_idx != 0 else LEFT
            
            if c_idx == 1:
                cell.fill = status_fill
                cell.font = BOLD_FONT
            else:
                cell.fill = row_fill

    auto_width(ws)

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--output', required=True)
    args = parser.parse_args()

    try:
        data = json.load(sys.stdin)
    except Exception as e:
        print(f"Failed to read JSON from stdin: {e}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.Workbook()

    if data.get("isAllEmployees"):
        # Master summary sheet
        ws_summary = wb.active
        ws_summary.title = "All Telecallers Summary"
        ws_summary.views.sheetView[0].showGridLines = True

        employees_data = data.get("employeesData", [])
        period_str = data.get("periodName", "Full Period")

        ws_summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        ws_summary.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

        t_cell = ws_summary.cell(row=1, column=1, value="ALL EMPLOYEES ATTENDANCE REPORT")
        t_cell.font = TITLE_FONT
        t_cell.fill = HEADER_FILL
        t_cell.alignment = CENTER

        subtitle = f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}  |  Period: {period_str}  |  Total Telecallers: {len(employees_data)}"
        s_cell = ws_summary.cell(row=2, column=1, value=subtitle)
        s_cell.font = Font(bold=False, color="FFFFFF", name="Calibri", size=10)
        s_cell.fill = ACCENT_FILL
        s_cell.alignment = CENTER

        ws_summary.row_dimensions[1].height = 30
        ws_summary.row_dimensions[2].height = 20

        headers = [
            "Employee Name", "Email", "Domain", "Branch", 
            "Workdays", "Present Days", "Absent Days", "Holidays", "Overtime Days", "Attendance Rate"
        ]

        for c_idx, h in enumerate(headers):
            cell = ws_summary.cell(row=4, column=c_idx+1, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER

        for idx, emp in enumerate(employees_data):
            r = 5 + idx
            s = emp.get("summary", {})
            pres = s.get("presentDays", 0)
            total_work = s.get("workingDays", 0)
            rate_num = round((pres / total_work) * 100, 1) if total_work > 0 else 0
            rate_str = f"{rate_num}%"

            vals = [
                emp.get("userName", "-"),
                emp.get("email", "-"),
                emp.get("domain", "-"),
                emp.get("branch", "-"),
                total_work,
                pres,
                s.get("absentDays", 0),
                s.get("holidays", 0),
                s.get("overtimeDays", 0),
                rate_str
            ]

            row_fill = ALT_ROW_FILL if idx % 2 == 1 else WHITE_FILL

            for c_idx, val in enumerate(vals):
                cell = ws_summary.cell(row=r, column=c_idx+1, value=val)
                cell.font = BOLD_FONT if c_idx == 0 or c_idx == 9 else BODY_FONT
                cell.border = THIN_BORDER
                cell.alignment = CENTER if c_idx >= 4 else LEFT
                cell.fill = row_fill

        auto_width(ws_summary)

        # Create individual worksheets for each employee
        used_titles = {"All Telecallers Summary"}
        for emp in employees_data:
            raw_title = emp.get("userName", "Employee")
            clean_title = re.sub(r'[\\/*?:\[\]]', '', raw_title)[:25].strip()
            if not clean_title: clean_title = "Emp"
            sheet_title = clean_title
            counter = 1
            while sheet_title in used_titles:
                sheet_title = f"{clean_title[:22]}_{counter}"
                counter += 1
            used_titles.add(sheet_title)

            ws_emp = wb.create_sheet(title=sheet_title)
            build_single_user_sheet(ws_emp, emp)

    else:
        ws = wb.active
        ws.title = "Attendance Report"
        build_single_user_sheet(ws, data)

    wb.save(args.output)
    print("Report generated successfully.")

if __name__ == "__main__":
    main()
