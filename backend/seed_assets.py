import openpyxl
import json
import sys
import os

def main():
    excel_path = "C:\\MyPers\\Projects\\GYCAnalysis\\asset.xlsx"
    if not os.path.exists(excel_path):
        print(json.dumps([]))
        return

    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active

        assets = []
        # Header is in row 1
        # Headers: ['Asset Photo', 'Asset Tag ID', 'Description', 'Brand', 'Status', 'Assigned to']
        for row in range(2, sheet.max_row + 1):
            tag_id = sheet.cell(row=row, column=2).value
            if not tag_id:
                continue
            
            photo = sheet.cell(row=row, column=1).value or ""
            desc = sheet.cell(row=row, column=3).value or ""
            brand = sheet.cell(row=row, column=4).value or ""
            status = sheet.cell(row=row, column=5).value or "Available"
            assigned_to = sheet.cell(row=row, column=6).value or ""
            
            assets.append({
                "assetPhoto": str(photo).strip(),
                "assetTagId": str(tag_id).strip(),
                "description": str(desc).strip(),
                "brand": str(brand).strip(),
                "status": str(status).strip(),
                "assignedTo": str(assigned_to).strip(),
                "assignedToName": str(assigned_to).strip()
            })
            
        print(json.dumps(assets))
    except Exception as e:
        sys.stderr.write(str(e) + "\n")
        sys.stderr.flush()
        print(json.dumps([]))

if __name__ == "__main__":
    main()
