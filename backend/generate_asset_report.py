import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json
import sys
import argparse
import os
from datetime import datetime

def parse_args():
    parser = argparse.ArgumentParser(description="Generate Asset Excel Report")
    parser.add_argument("--output", required=True, help="Path to save output xlsx")
    parser.add_argument("--year", help="Year filter (e.g. 2026)")
    parser.add_argument("--month", help="Month filter (e.g. 06)")
    return parser.parse_args()

def matches_filter(date_str, year, month):
    if not date_str:
        return False
    # date_str can be YYYY-MM (like '2026-06') or ISO datetime (like '2026-06-17T12:00:00Z')
    try:
        if 'T' in date_str:
            dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
            y_val = str(dt.year)
            m_val = f"{dt.month:02d}"
        else:
            parts = date_str.split('-')
            y_val = parts[0]
            m_val = parts[1] if len(parts) > 1 else ""
        
        if year and y_val != year:
            return False
        if month and m_val != month:
            return False
        return True
    except Exception:
        # Fallback string checks if formatting fails
        if year and year not in date_str:
            return False
        if month and f"-{month}" not in date_str:
            return False
        return True

def main():
    args = parse_args()

    # Read data from stdin
    try:
        raw_input = sys.stdin.read()
        if not raw_input.strip():
            raw_input = "{}"
        data = json.loads(raw_input)
    except Exception as e:
        sys.stderr.write(f"Error reading JSON from stdin: {str(e)}\n")
        sys.stderr.flush()
        data = {}

    assets = data.get("assets", [])
    verifications = data.get("verifications", [])
    notifications = data.get("notifications", [])

    # Filter verifications and notifications
    if args.year or args.month:
        verifications = [v for v in verifications if matches_filter(v.get("month") or v.get("submittedAt"), args.year, args.month)]
        notifications = [n for n in notifications if matches_filter(n.get("createdAt"), args.year, args.month)]

    # Create workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="111111", end_color="111111", fill_type="solid")
    cell_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    # --- Sheet 1: Assets ---
    ws_assets = wb.create_sheet(title="Assets")
    ws_assets.views.sheetView[0].showGridLines = True
    
    # Headers
    headers_assets = ['Asset Photo', 'Asset Tag ID', 'Description', 'Brand', 'Status', 'Assigned to']
    ws_assets.append(headers_assets)
    
    for asset in assets:
        row_data = [
            asset.get("assetPhoto") or "",
            asset.get("assetTagId") or "",
            asset.get("description") or "",
            asset.get("brand") or "",
            asset.get("status") or "Available",
            asset.get("assignedTo") or ""
        ]
        ws_assets.append(row_data)

    # --- Sheet 2: Verifications ---
    ws_ver = wb.create_sheet(title="Verifications")
    ws_ver.views.sheetView[0].showGridLines = True
    
    headers_ver = ['Verification ID', 'User Name', 'User Email', 'Month', 'Assets Declared', 'Has Issues', 'Repaired Handed Over', 'New Device Received', 'New Asset ID', 'Submitted At']
    ws_ver.append(headers_ver)
    
    for v in verifications:
        # Format declared assets list
        decl_assets = v.get("assets", [])
        decl_text_list = []
        for da in decl_assets:
            t = da.get("type", "")
            c = da.get("code") or ""
            p = da.get("phoneNumber") or ""
            prov = da.get("provider") or ""
            if t == "SIM":
                decl_text_list.append(f"SIM: {p} ({prov})")
            else:
                decl_text_list.append(f"{t}: {c}")
        decl_text = ", ".join(decl_text_list) if decl_text_list else "None"

        row_data = [
            v.get("id") or "",
            v.get("name") or "",
            v.get("email") or "",
            v.get("month") or "",
            decl_text,
            "Yes" if v.get("hasIssues") else "No",
            "Yes" if v.get("repairedHandedOver") else ("No" if v.get("repairedHandedOver") == False else "N/A"),
            "Yes" if v.get("newDeviceReceived") else ("No" if v.get("newDeviceReceived") == False else "N/A"),
            v.get("newAssetTagId") or "",
            v.get("submittedAt") or ""
        ]
        ws_ver.append(row_data)

    # --- Sheet 3: Notifications ---
    ws_notif = wb.create_sheet(title="Admin Alerts")
    ws_notif.views.sheetView[0].showGridLines = True
    
    headers_notif = ['Alert ID', 'User Name', 'User Email', 'Alert Type', 'Message', 'Resolved', 'Created At']
    ws_notif.append(headers_notif)
    
    for n in notifications:
        row_data = [
            n.get("id") or "",
            n.get("userName") or "",
            n.get("userEmail") or "",
            n.get("type") or "",
            n.get("message") or "",
            "Yes" if n.get("resolved") else "No",
            n.get("createdAt") or ""
        ]
        ws_notif.append(row_data)

    # Format sheets
    for ws in [ws_assets, ws_ver, ws_notif]:
        # Format Header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
        
        # Format Data rows and auto-adjust widths
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = align_left
        
        # Set row heights
        ws.row_dimensions[1].height = 28
        for r in range(2, ws.max_row + 1):
            ws.row_dimensions[r].height = 20

        # Adjust column widths dynamically
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            # Add padding
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Save to output path
    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    wb.save(args.output)
    print(f"Report written successfully to {args.output}")

if __name__ == "__main__":
    main()
